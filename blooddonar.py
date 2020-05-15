#python code to make or udate data of user for blood donation
#hackathon 1
#Team Member : PRIYAM (19BCS089) and SATYAM KUMAR (19BEC040)
# exact address of the file has to be placed in first argument of Path

import openpyxl
from pathlib import Path

# Donar file
xlsx_file = Path("/home/iiit/Documents/data structure programs","donar.xlsx")
wb_obj = openpyxl.load_workbook(xlsx_file)
sheet = wb_obj.active

# Reciepient file
xlsx_file1 = Path("/home/iiit/Documents/data structure programs","reciepient.xlsx")
wb_obj1 = openpyxl.load_workbook(xlsx_file1)
sheet1 = wb_obj1.active

d=[]
reciepient=[] 
i = 1
j = 1
print('\n1. Enter Donor data\n2. Enter Recipient data\n3. Get Donor data \
\n4. Get Recipient data\n5. Exit From The system')
ch = int(input('\nEnter your choice : '))

temp = sheet.cell(row = 3,column = 3)
temp.value = "abc"
if ch==1:
    print('\nEnter the following details of the Donar: ')
    d.append = input("\nDonar ID: ")
    d.append = input("\nName: ")
    d.append = input("\nBloodgroup: ")
    d.append = input("\nAge: ")
    d.append = input("\nGender: ")
    d.append = input("\nPhone Number: ")
    d.append = input("\nEmail ID: ")
    d.append = input("\nAddress: ")
    for data in d:
        temp = sheet.cell(row = 2,column = 2)
        temp.value = data
if ch==2:
    temp = sheet.cell(row = 3,column = 3)
    temp.value = "abc"
wb_obj.save("/home/iiit/Documents/data structure programs","donar.xlsx")